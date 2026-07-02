"""将平台预设四级导航树 JSON 导出为 Excel。"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "层级",
    "层级名称",
    "节点键",
    "父节点键",
    "模块ID",
    "模块名称",
    "一级导航",
    "二级导航",
    "三级分组",
    "四级设置项",
    "标题",
    "英文标题",
    "路径",
    "groupKey",
    "seq",
    "功能模块",
    "功能表述",
    "功能说明",
    "featureId",
    "树状路径",
]

L4_HEADERS = [
    "一级导航",
    "二级导航",
    "三级分组",
    "四级设置项",
    "seq",
    "节点键",
    "路径",
    "groupKey",
    "功能模块",
    "功能表述",
    "功能说明",
]


def style_header_row(ws, row_num=1):
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[row_num]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize_columns(ws, max_width=48, wide_cols=None):
    wide_cols = wide_cols or set()
    for col_idx, column_cells in enumerate(ws.columns, 1):
        letter = get_column_letter(col_idx)
        length = max(len(str(cell.value or "")) for cell in column_cells)
        cap = 72 if col_idx in wide_cols else max_width
        ws.column_dimensions[letter].width = min(max(length + 2, 10), cap)


def row_values(row):
    return [
        row["level"],
        row["levelLabel"],
        row["key"],
        row["parentKey"],
        row["moduleId"],
        row["moduleTitle"],
        row["l1Title"],
        row["l2Title"],
        row["l3Title"],
        row["l4Title"],
        row["title"],
        row["titleEn"],
        row["path"],
        row["groupKey"],
        row["seq"],
        row.get("moduleName", ""),
        row.get("sceneDesc", ""),
        row.get("feature", ""),
        row["featureId"],
        row["treePath"],
    ]


def l4_row_values(row):
    return [
        row["l1Title"],
        row["l2Title"],
        row["l3Title"],
        row["l4Title"],
        row["seq"],
        row["key"],
        row["path"],
        row["groupKey"],
        row.get("moduleName", ""),
        row.get("sceneDesc", ""),
        row.get("feature", ""),
    ]


def write_sheet_tree(wb, rows):
    ws = wb.create_sheet("树状结构")
    ws.append(HEADERS)
    style_header_row(ws)

    for row in rows:
        ws.append(row_values(row))

    ws.freeze_panes = "A2"
    # 功能表述(17)、功能说明(18)
    autosize_columns(ws, wide_cols={17, 18})
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (17, 18, 20))


def write_sheet_l4_settings(wb, rows):
    ws = wb.create_sheet("分组内功能设置")
    ws.append(L4_HEADERS)
    style_header_row(ws)

    l4_rows = [r for r in rows if r["level"] == 4]
    for row in l4_rows:
        ws.append(l4_row_values(row))

    ws.freeze_panes = "A2"
    # 功能表述(10)、功能说明(11)
    autosize_columns(ws, wide_cols={10, 11})
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (10, 11))


def write_sheet_stats(wb, stats):
    ws = wb.create_sheet("统计", 0)
    ws.append(["指标", "数量"])
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
    items = [
        ("一级导航（模块）", stats["moduleCount"]),
        ("二级导航", stats["l2Count"]),
        ("三级分组", stats["l3Count"]),
        ("四级设置项", stats["l4Count"]),
        ("节点合计", stats["totalNodes"]),
    ]
    for label, count in items:
        ws.append([label, count])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12


def write_sheet_by_level(wb, rows):
    ws = wb.create_sheet("按层级汇总")
    ws.append(["层级", "层级名称", "节点键", "标题", "模块名称", "路径", "groupKey", "seq", "功能表述"])
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
    for row in rows:
        ws.append([
            row["level"],
            row["levelLabel"],
            row["key"],
            row["title"],
            row["moduleTitle"],
            row["path"],
            row["groupKey"],
            row["seq"],
            row.get("sceneDesc", ""),
        ])
    ws.freeze_panes = "A2"
    autosize_columns(ws, wide_cols={9})


def main():
    if len(sys.argv) < 3:
        print("Usage: python platform-preset-tree-xlsx.py <input.json> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    data = json.loads(input_path.read_text(encoding="utf-8"))
    rows = data["rows"]
    stats = data["stats"]

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet_stats(wb, stats)
    write_sheet_tree(wb, rows)
    write_sheet_l4_settings(wb, rows)
    write_sheet_by_level(wb, rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
