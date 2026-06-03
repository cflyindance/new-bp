# -*- coding: utf-8 -*-
"""Export 设置项功能归类分析-索引.md table to Excel."""
import re
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    import sys

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
MD_PATH = ROOT / "设置项功能归类分析-索引.md"
XLSX_PATH = ROOT / "设置项功能归类分析-索引.xlsx"

HEADERS = [
    "日期",
    "原始一级导航",
    "原始二级导航",
    "原始功能",
    "功能",
    "功能描述（使用场景）",
    "功能涉及的产品线",
    "结论（一级导航）",
    "一级导航下归属分类",
    "为何归入该一级导航（简述）",
]


def clean_cell(text: str) -> str:
    text = text.strip()
    text = re.sub(r"\*\*", "", text)
    text = re.sub(r"`([^`]+)`", r"\1", text)
    return text


def parse_rows():
    lines = MD_PATH.read_text(encoding="utf-8").splitlines()
    rows = []
    for line in lines:
        if not re.match(r"^\| 20\d{2}-", line):
            continue
        parts = [p.strip() for p in line.split("|")]
        # leading/trailing empty from split
        if parts and parts[0] == "":
            parts = parts[1:]
        if parts and parts[-1] == "":
            parts = parts[:-1]
        if len(parts) < 10:
            continue
        rows.append([clean_cell(c) for c in parts[:10]])
    return rows


def autosize_columns(ws):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 80))
        ws.column_dimensions[letter].width = max(10, min(max_len + 2, 60))


def main():
    data = parse_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "设置项索引"

    header_font = Font(bold=True)
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    wrap = Alignment(wrap_text=True, vertical="top")
    for r_idx, row in enumerate(data, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = wrap

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(data) + 1}"
    autosize_columns(ws)

    wb.save(XLSX_PATH)
    print(f"Exported {len(data)} rows -> {XLSX_PATH}")


if __name__ == "__main__":
    main()
